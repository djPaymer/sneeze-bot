import logging
import io
import os
from datetime import datetime, date, timedelta, timezone
from typing import Optional
from telegram import Update
from telegram.ext import Application, CommandHandler, MessageHandler, ContextTypes, filters
from telegram import ReplyKeyboardMarkup, KeyboardButton
import matplotlib
matplotlib.use('Agg')  # Используем backend без GUI
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib import font_manager
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import config
from database import Database


def get_utc_today():
    """Получает сегодняшнюю дату в UTC"""
    return datetime.now(timezone.utc).date()


def get_user_date_from_message(update: Update):
    """Получает дату пользователя из сообщения (учитывает часовой пояс)"""
    if update.message and update.message.date:
        # Дата сообщения уже в UTC, конвертируем в date
        return update.message.date.date()
    return get_utc_today()


def is_admin(user_id: int) -> bool:
    """Проверяет, является ли пользователь администратором"""
    return user_id in config.ADMIN_IDS


def create_excel_export(start_date: Optional[str] = None, end_date: Optional[str] = None) -> io.BytesIO:
    """
    Создает Excel файл с экспортом статистики всех пользователей
    
    Args:
        start_date: Начальная дата в формате YYYY-MM-DD (если None, все записи)
        end_date: Конечная дата в формате YYYY-MM-DD (если None, все записи)
    
    Returns:
        BytesIO объект с Excel файлом
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Статистика чиханий"
    
    # Стили
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14)
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    # Заголовок
    ws['A1'] = "Статистика чиханий"
    ws['A1'].font = title_font
    if start_date and end_date:
        ws['A2'] = f"Период: {start_date} - {end_date}"
    else:
        ws['A2'] = "За весь период"
    ws.merge_cells('A1:D1')
    ws.merge_cells('A2:D2')
    
    # Получаем данные
    all_stats = db.get_all_users_stats(start_date, end_date)
    detailed_stats = db.get_all_users_detailed_stats(start_date, end_date)
    
    # Лист 1: Общая статистика по пользователям
    row = 4
    ws.cell(row=row, column=1, value="User ID").fill = header_fill
    ws.cell(row=row, column=1).font = header_font
    ws.cell(row=row, column=1).alignment = center_alignment
    ws.cell(row=row, column=2, value="Всего чиханий").fill = header_fill
    ws.cell(row=row, column=2).font = header_font
    ws.cell(row=row, column=2).alignment = center_alignment
    
    row += 1
    total_all = 0
    for user_id, total in all_stats:
        ws.cell(row=row, column=1, value=user_id)
        ws.cell(row=row, column=2, value=total)
        total_all += total
        row += 1
    
    # Итого
    ws.cell(row=row, column=1, value="ИТОГО").font = Font(bold=True)
    ws.cell(row=row, column=2, value=total_all).font = Font(bold=True)
    
    # Автоподбор ширины колонок
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    
    # Лист 2: Детальная статистика
    ws2 = wb.create_sheet(title="Детальная статистика")
    
    row = 1
    ws2.cell(row=row, column=1, value="User ID").fill = header_fill
    ws2.cell(row=row, column=1).font = header_font
    ws2.cell(row=row, column=1).alignment = center_alignment
    ws2.cell(row=row, column=2, value="Дата").fill = header_fill
    ws2.cell(row=row, column=2).font = header_font
    ws2.cell(row=row, column=2).alignment = center_alignment
    ws2.cell(row=row, column=3, value="Количество").fill = header_fill
    ws2.cell(row=row, column=3).font = header_font
    ws2.cell(row=row, column=3).alignment = center_alignment
    
    row += 1
    for user_id, date_str, count in detailed_stats:
        ws2.cell(row=row, column=1, value=user_id)
        ws2.cell(row=row, column=2, value=date_str)
        ws2.cell(row=row, column=3, value=count)
        row += 1
    
    # Автоподбор ширины колонок
    ws2.column_dimensions['A'].width = 15
    ws2.column_dimensions['B'].width = 15
    ws2.column_dimensions['C'].width = 15
    
    # Сохраняем в BytesIO
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer


# Настройка логирования
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# Инициализация базы данных
db = Database()


def format_stats(stats: list, period_title: str) -> str:
    """Форматирует статистику в читаемый вид"""
    if not stats:
        return f"За период {period_title} записей нет."
    
    total = sum(count for _, count in stats)
    avg = total / len(stats) if stats else 0
    
    result = f"📊 Статистика за {period_title}:\n\n"
    result += f"Всего дней с записями: {len(stats)}\n"
    result += f"Общее количество чиханий: {total}\n"
    result += f"Среднее за день: {avg:.1f}\n\n"
    result += "Детализация по дням:\n"
    
    for day_date, count in stats:
        # Преобразуем YYYY-MM-DD в DD.MM
        day_parts = day_date.split('-')
        day_formatted = f"{day_parts[2]}.{day_parts[1]}"
        result += f"  {day_formatted}: {count} раз\n"
    
    return result


def create_stats_chart(stats: list, title: str) -> io.BytesIO:
    """Создает график статистики чиханий"""
    if not stats:
        return None
    
    # Подготовка данных
    dates = []
    counts = []
    
    for day_date, count in stats:
        # Преобразуем YYYY-MM-DD в объект date
        date_obj = datetime.strptime(day_date, '%Y-%m-%d').date()
        dates.append(date_obj)
        counts.append(count)
    
    # Создание графика
    plt.figure(figsize=(12, 6))
    plt.plot(dates, counts, marker='o', linewidth=2, markersize=8, color='#4CAF50')
    plt.fill_between(dates, counts, alpha=0.3, color='#4CAF50')
    
    # Настройка осей
    plt.xlabel('Дата', fontsize=12, fontweight='bold')
    plt.ylabel('Количество чиханий', fontsize=12, fontweight='bold')
    plt.title(f'Статистика чиханий: {title}', fontsize=14, fontweight='bold', pad=20)
    
    # Форматирование дат на оси X
    plt.gca().xaxis.set_major_formatter(mdates.DateFormatter('%d.%m'))
    # Для недели показываем все дни, для месяца - интервалы
    if len(dates) <= 7:
        plt.gca().xaxis.set_major_locator(mdates.DayLocator(interval=1))
    else:
        plt.gca().xaxis.set_major_locator(mdates.DayLocator(interval=max(1, len(dates) // 10)))
    plt.xticks(rotation=45, ha='right')
    
    # Сетка
    plt.grid(True, alpha=0.3, linestyle='--')
    
    # Улучшение внешнего вида
    plt.tight_layout()
    
    # Сохранение в BytesIO
    buf = io.BytesIO()
    plt.savefig(buf, format='png', dpi=100, bbox_inches='tight')
    buf.seek(0)
    plt.close()
    
    return buf


def get_reply_keyboard():
    """Создает постоянную клавиатуру с кнопками"""
    keyboard = [
        [KeyboardButton("🤧 Чихнуть")],
        [KeyboardButton("📊 Статистика"), KeyboardButton("📈 График")],
        [KeyboardButton("📋 Меню")]
    ]
    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True)


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик команды /start"""
    user_id = update.effective_user.id
    welcome_message = (
        "👋 Привет! Я бот для отслеживания чиханий.\n\n"
        "Доступные команды:\n"
        "/add <количество> - добавить количество чиханий за сегодня\n"
        "/stats - статистика за текущую неделю (текст)\n"
        "/stats week - статистика за неделю\n"
        "/stats month - статистика за текущий месяц\n"
        "/stats <месяц> <год> - статистика за конкретный месяц (например: /stats 12 2024)\n"
        "/stats <дата1> <дата2> - статистика за период (формат: ДД.ММ.ГГГГ)\n"
        "/chart - график за текущую неделю\n"
        "/chart week/month/<месяц> <год>/<дата1> <дата2> - график за период\n"
        "/edit <дата> <количество> - редактировать данные за дату (формат: ДД.ММ.ГГГГ)\n"
        "/today - посмотреть количество чиханий за сегодня\n"
    )
    
    # Добавляем админ-команды, если пользователь администратор
    if is_admin(user_id):
        welcome_message += (
            "\n🔐 Админ-команды:\n"
            "/admin_stats - статистика всех пользователей\n"
            "/admin_stats <дата1> <дата2> - статистика за период (формат: ДД.ММ.ГГГГ)\n"
            "/admin_export - экспорт в Excel (все данные)\n"
            "/admin_export <дата1> <дата2> - экспорт за период (формат: ДД.ММ.ГГГГ)\n"
        )
    
    welcome_message += (
        "\n"
        "Также вы можете просто написать число - оно будет записано как количество чиханий за сегодня.\n\n"
        "Или используйте кнопки внизу экрана:"
    )
    
    await update.message.reply_text(
        welcome_message,
        reply_markup=get_reply_keyboard()
    )


async def add_sneeze(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик команды /add"""
    user_id = update.effective_user.id
    
    if not context.args:
        await update.message.reply_text(
            "❌ Пожалуйста, укажите количество чиханий.\n"
            "Пример: /add 5"
        )
        return
    
    try:
        count = int(context.args[0])
        if count < 0:
            await update.message.reply_text("❌ Количество не может быть отрицательным!")
            return
        
        # Используем дату из сообщения пользователя для корректной работы с часовыми поясами
        user_today = get_user_date_from_message(update)
        today = user_today.isoformat()
        db.add_sneeze(user_id, count, today)
        
        await update.message.reply_text(
            f"✅ Записано: {count} чиханий за сегодня ({user_today.strftime('%d.%m.%Y')})\n"
            f"🤧 Будь здоров!",
            reply_markup=get_reply_keyboard()
        )
    except ValueError:
        await update.message.reply_text("❌ Пожалуйста, введите число!")


async def show_stats(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик команды /stats"""
    user_id = update.effective_user.id
    now = datetime.now(timezone.utc)
    
    # Используем дату пользователя для корректного отображения
    user_today = get_user_date_from_message(update)
    
    # Проверяем наличие args и их количество
    args = context.args if context.args is not None else []
    
    stats = None
    period_title = ""
    
    if len(args) == 0:
        # По умолчанию - статистика за текущую неделю
        stats = db.get_week_stats(user_id, user_today.isoformat())
        week_start = user_today - timedelta(days=6)
        period_title = f"неделю ({week_start.strftime('%d.%m')} - {user_today.strftime('%d.%m.%Y')})"
        
    elif len(args) == 1:
        # /stats week или /stats month
        if args[0].lower() == 'week':
            stats = db.get_week_stats(user_id, user_today.isoformat())
            week_start = user_today - timedelta(days=6)
            period_title = f"неделю ({week_start.strftime('%d.%m')} - {user_today.strftime('%d.%m.%Y')})"
        elif args[0].lower() == 'month':
            year = now.year
            month = now.month
            stats = db.get_month_stats(user_id, year, month)
            month_names = ['Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь',
                          'Июль', 'Август', 'Сентябрь', 'Октябрь', 'Ноябрь', 'Декабрь']
            period_title = f"{month_names[month - 1]} {year}"
        else:
            await update.message.reply_text(
                "❌ Неверный формат. Используйте: /stats, /stats week, /stats month или /stats <месяц> <год>",
                reply_markup=get_reply_keyboard()
            )
            return
            
    elif len(args) == 2:
        # Проверяем, это даты или месяц/год
        try:
            # Пробуем распарсить как даты (ДД.ММ.ГГГГ)
            date1_parts = args[0].split('.')
            date2_parts = args[1].split('.')
            
            if len(date1_parts) == 3 and len(date2_parts) == 3:
                # Это даты - период
                day1, month1, year1 = map(int, date1_parts)
                day2, month2, year2 = map(int, date2_parts)
                start_date = date(year1, month1, day1)
                end_date = date(year2, month2, day2) + timedelta(days=1)  # +1 чтобы включить последний день
                
                if start_date > end_date:
                    await update.message.reply_text(
                        "❌ Начальная дата должна быть раньше конечной!",
                        reply_markup=get_reply_keyboard()
                    )
                    return
                
                stats = db.get_period_stats(user_id, start_date.isoformat(), end_date.isoformat())
                period_title = f"период ({args[0]} - {args[1]})"
            else:
                # Это месяц и год
                month = int(args[0])
                year = int(args[1])
                if month < 1 or month > 12:
                    await update.message.reply_text("❌ Месяц должен быть от 1 до 12!", reply_markup=get_reply_keyboard())
                    return
                stats = db.get_month_stats(user_id, year, month)
                month_names = ['Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь',
                              'Июль', 'Август', 'Сентябрь', 'Октябрь', 'Ноябрь', 'Декабрь']
                period_title = f"{month_names[month - 1]} {year}"
        except ValueError:
            await update.message.reply_text(
                "❌ Неверный формат. Используйте:\n"
                "/stats - за неделю\n"
                "/stats week - за неделю\n"
                "/stats month - за месяц\n"
                "/stats <месяц> <год> - за конкретный месяц\n"
                "/stats <дата1> <дата2> - за период (формат: ДД.ММ.ГГГГ)",
                reply_markup=get_reply_keyboard()
            )
            return
    else:
        await update.message.reply_text(
            "❌ Неверный формат команды.\n"
            "Используйте:\n"
            "/stats - за неделю\n"
            "/stats week - за неделю\n"
            "/stats month - за месяц\n"
            "/stats <месяц> <год> - за конкретный месяц\n"
            "/stats <дата1> <дата2> - за период (формат: ДД.ММ.ГГГГ)",
            reply_markup=get_reply_keyboard()
        )
        return
    
    message = format_stats(stats, period_title)
    
    # Отправляем только текстовую статистику (без графика)
    await update.message.reply_text(message, reply_markup=get_reply_keyboard())
    
    # Сохраняем данные статистики в context для использования в графике
    context.user_data['last_stats'] = stats
    context.user_data['last_period_title'] = period_title


async def show_chart(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик команды /chart - показывает график для последнего запрошенного периода"""
    user_id = update.effective_user.id
    now = datetime.now(timezone.utc)
    
    # Используем дату пользователя для корректного отображения
    user_today = get_user_date_from_message(update)
    
    # Проверяем наличие args и их количество
    args = context.args if context.args is not None else []
    
    stats = None
    period_title = ""
    
    if len(args) == 0:
        # По умолчанию - график за текущую неделю
        stats = db.get_week_stats(user_id, user_today.isoformat())
        week_start = user_today - timedelta(days=6)
        period_title = f"неделю ({week_start.strftime('%d.%m')} - {user_today.strftime('%d.%m.%Y')})"
        
    elif len(args) == 1:
        # /chart week или /chart month
        if args[0].lower() == 'week':
            stats = db.get_week_stats(user_id, user_today.isoformat())
            week_start = user_today - timedelta(days=6)
            period_title = f"неделю ({week_start.strftime('%d.%m')} - {user_today.strftime('%d.%m.%Y')})"
        elif args[0].lower() == 'month':
            year = user_today.year
            month = user_today.month
            stats = db.get_month_stats(user_id, year, month)
            month_names = ['Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь',
                          'Июль', 'Август', 'Сентябрь', 'Октябрь', 'Ноябрь', 'Декабрь']
            period_title = f"{month_names[month - 1]} {year}"
        else:
            await update.message.reply_text(
                "❌ Неверный формат. Используйте: /chart, /chart week, /chart month или /chart <месяц> <год>",
                reply_markup=get_reply_keyboard()
            )
            return
            
    elif len(args) == 2:
        # Проверяем, это даты или месяц/год
        try:
            # Пробуем распарсить как даты (ДД.ММ.ГГГГ)
            date1_parts = args[0].split('.')
            date2_parts = args[1].split('.')
            
            if len(date1_parts) == 3 and len(date2_parts) == 3:
                # Это даты - период
                day1, month1, year1 = map(int, date1_parts)
                day2, month2, year2 = map(int, date2_parts)
                start_date = date(year1, month1, day1)
                end_date = date(year2, month2, day2) + timedelta(days=1)
                
                if start_date > end_date:
                    await update.message.reply_text(
                        "❌ Начальная дата должна быть раньше конечной!",
                        reply_markup=get_reply_keyboard()
                    )
                    return
                
                stats = db.get_period_stats(user_id, start_date.isoformat(), end_date.isoformat())
                period_title = f"период ({args[0]} - {args[1]})"
            else:
                # Это месяц и год
                month = int(args[0])
                year = int(args[1])
                if month < 1 or month > 12:
                    await update.message.reply_text("❌ Месяц должен быть от 1 до 12!", reply_markup=get_reply_keyboard())
                    return
                stats = db.get_month_stats(user_id, year, month)
                month_names = ['Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь',
                              'Июль', 'Август', 'Сентябрь', 'Октябрь', 'Ноябрь', 'Декабрь']
                period_title = f"{month_names[month - 1]} {year}"
        except ValueError:
            await update.message.reply_text(
                "❌ Неверный формат. Используйте:\n"
                "/chart - за неделю\n"
                "/chart week - за неделю\n"
                "/chart month - за месяц\n"
                "/chart <месяц> <год> - за конкретный месяц\n"
                "/chart <дата1> <дата2> - за период (формат: ДД.ММ.ГГГГ)",
                reply_markup=get_reply_keyboard()
            )
            return
    else:
        await update.message.reply_text(
            "❌ Неверный формат команды.\n"
            "Используйте:\n"
            "/chart - за неделю\n"
            "/chart week - за неделю\n"
            "/chart month - за месяц\n"
            "/chart <месяц> <год> - за конкретный месяц\n"
            "/chart <дата1> <дата2> - за период (формат: ДД.ММ.ГГГГ)",
            reply_markup=get_reply_keyboard()
        )
        return
    
    # Отправляем график
    if stats:
        chart_buffer = create_stats_chart(stats, period_title)
        if chart_buffer:
            await update.message.reply_photo(
                photo=chart_buffer,
                caption=f"📈 График чиханий за {period_title}",
                reply_markup=get_reply_keyboard()
            )
        else:
            await update.message.reply_text(
                f"❌ Не удалось создать график за {period_title}",
                reply_markup=get_reply_keyboard()
            )
    else:
        await update.message.reply_text(
            f"❌ Нет данных за {period_title} для построения графика",
            reply_markup=get_reply_keyboard()
        )


async def edit_date(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик команды /edit"""
    user_id = update.effective_user.id
    
    if len(context.args) != 2:
        await update.message.reply_text(
            "❌ Неверный формат команды.\n"
            "Используйте: /edit <дата> <количество>\n"
            "Формат даты: ДД.ММ.ГГГГ\n"
            "Пример: /edit 15.12.2024 10"
        )
        return
    
    try:
        # Парсим дату из формата ДД.ММ.ГГГГ
        date_str = context.args[0]
        date_parts = date_str.split('.')
        if len(date_parts) != 3:
            raise ValueError("Неверный формат даты")
        
        day, month, year = map(int, date_parts)
        target_date = date(year, month, day).isoformat()
        
        count = int(context.args[1])
        if count < 0:
            await update.message.reply_text("❌ Количество не может быть отрицательным!")
            return
        
        db.update_date_count(user_id, target_date, count)
        
        await update.message.reply_text(
            f"✅ Обновлено: {count} чиханий за {date_str}"
        )
    except ValueError as e:
        await update.message.reply_text(
            f"❌ Ошибка в формате данных!\n"
            f"Используйте: /edit ДД.ММ.ГГГГ <количество>\n"
            f"Пример: /edit 15.12.2024 10"
        )
    except Exception as e:
        await update.message.reply_text(f"❌ Ошибка: {str(e)}")


async def show_today(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик команды /today"""
    user_id = update.effective_user.id
    user_today = get_user_date_from_message(update)
    today = user_today.isoformat()
    
    count = db.get_date_count(user_id, today)
    
    if count is None:
        await update.message.reply_text(
            f"📅 За сегодня ({user_today.strftime('%d.%m.%Y')}) записей нет.\n"
            f"Используйте /add <количество> или просто напишите число.",
            reply_markup=get_reply_keyboard()
        )
    else:
        await update.message.reply_text(
            f"📅 Сегодня ({user_today.strftime('%d.%m.%Y')}): {count} чиханий",
            reply_markup=get_reply_keyboard()
        )


async def handle_number_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик сообщений с числами и кнопок"""
    user_id = update.effective_user.id
    text = update.message.text.strip()
    
    # Проверяем, нажата ли кнопка "Чихнуть"
    if text == "🤧 Чихнуть":
        user_today = get_user_date_from_message(update)
        today = user_today.isoformat()
        new_count = db.increment_sneeze(user_id, today)
        
        if new_count is not None:
            await update.message.reply_text(
                f"✅ Записано чихание!\n"
                f"📊 Сегодня ({user_today.strftime('%d.%m.%Y')}): {new_count} чиханий\n"
                f"🤧 Будь здоров!",
                reply_markup=get_reply_keyboard()
            )
        else:
            await update.message.reply_text(
                "❌ Ошибка при записи",
                reply_markup=get_reply_keyboard()
            )
        return
    
    # Проверяем, нажата ли кнопка "Статистика"
    if text == "📊 Статистика":
        await show_stats(update, context)
        return
    
    # Проверяем, нажата ли кнопка "График"
    if text == "📈 График":
        await show_chart(update, context)
        return
    
    # Проверяем, нажата ли кнопка "Меню"
    if text == "📋 Меню":
        await start(update, context)
        return
    
    # Обработка чисел
    try:
        count = int(text)
        if count < 0:
            await update.message.reply_text(
                "❌ Количество не может быть отрицательным!",
                reply_markup=get_reply_keyboard()
            )
            return
        
        user_today = get_user_date_from_message(update)
        today = user_today.isoformat()
        db.add_sneeze(user_id, count, today)
        
        await update.message.reply_text(
            f"✅ Записано: {count} чиханий за сегодня ({user_today.strftime('%d.%m.%Y')})\n"
            f"🤧 Будь здоров!",
            reply_markup=get_reply_keyboard()
        )
    except ValueError:
        # Если это не число и не кнопка, игнорируем
        pass


async def admin_stats(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик команды /admin_stats - статистика всех пользователей (только для админов)"""
    user_id = update.effective_user.id
    
    if not is_admin(user_id):
        await update.message.reply_text("❌ У вас нет прав администратора!")
        return
    
    args = context.args if context.args is not None else []
    start_date = None
    end_date = None
    
    # Парсинг аргументов для периода
    if len(args) == 2:
        try:
            # Формат: ДД.ММ.ГГГГ
            date1_parts = args[0].split('.')
            date2_parts = args[1].split('.')
            
            if len(date1_parts) == 3 and len(date2_parts) == 3:
                day1, month1, year1 = map(int, date1_parts)
                day2, month2, year2 = map(int, date2_parts)
                start_date = date(year1, month1, day1).isoformat()
                end_date = (date(year2, month2, day2) + timedelta(days=1)).isoformat()
        except ValueError:
            await update.message.reply_text(
                "❌ Неверный формат дат. Используйте: /admin_stats или /admin_stats ДД.ММ.ГГГГ ДД.ММ.ГГГГ"
            )
            return
    
    # Получаем статистику
    all_stats = db.get_all_users_stats(start_date, end_date)
    
    if not all_stats:
        period_text = f" за период {args[0]} - {args[1]}" if start_date and end_date else ""
        await update.message.reply_text(f"📊 Нет данных{period_text}.")
        return
    
    # Формируем сообщение
    period_text = f" за период {args[0]} - {args[1]}" if start_date and end_date else " (за весь период)"
    message = f"📊 Статистика всех пользователей{period_text}:\n\n"
    
    total_all = 0
    for user_id_stat, total in all_stats:
        message += f"👤 User ID: {user_id_stat}\n"
        message += f"   Всего чиханий: {total}\n\n"
        total_all += total
    
    message += f"━━━━━━━━━━━━━━━━\n"
    message += f"📈 ИТОГО: {total_all} чиханий\n"
    message += f"👥 Пользователей: {len(all_stats)}"
    
    await update.message.reply_text(message)


async def admin_export(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик команды /admin_export - экспорт в Excel (только для админов)"""
    user_id = update.effective_user.id
    
    if not is_admin(user_id):
        await update.message.reply_text("❌ У вас нет прав администратора!")
        return
    
    args = context.args if context.args is not None else []
    start_date = None
    end_date = None
    
    # Парсинг аргументов для периода
    if len(args) == 2:
        try:
            # Формат: ДД.ММ.ГГГГ
            date1_parts = args[0].split('.')
            date2_parts = args[1].split('.')
            
            if len(date1_parts) == 3 and len(date2_parts) == 3:
                day1, month1, year1 = map(int, date1_parts)
                day2, month2, year2 = map(int, date2_parts)
                start_date = date(year1, month1, day1).isoformat()
                end_date = (date(year2, month2, day2) + timedelta(days=1)).isoformat()
        except ValueError:
            await update.message.reply_text(
                "❌ Неверный формат дат. Используйте: /admin_export или /admin_export ДД.ММ.ГГГГ ДД.ММ.ГГГГ"
            )
            return
    
    try:
        await update.message.reply_text("⏳ Создаю Excel файл...")
        
        # Создаем Excel файл
        excel_buffer = create_excel_export(start_date, end_date)
        
        # Формируем имя файла
        if start_date and end_date:
            filename = f"sneeze_stats_{start_date}_to_{end_date.replace('-', '')[:8]}.xlsx"
        else:
            filename = f"sneeze_stats_all_{datetime.now().strftime('%Y%m%d')}.xlsx"
        
        # Отправляем файл
        await update.message.reply_document(
            document=excel_buffer,
            filename=filename,
            caption="📊 Экспорт статистики чиханий"
        )
    except Exception as e:
        logger.error(f"Ошибка при создании Excel файла: {e}")
        await update.message.reply_text(f"❌ Ошибка при создании Excel файла: {str(e)}")


def main():
    """Запуск бота"""
    if not config.BOT_TOKEN:
        logger.error("BOT_TOKEN не установлен! Создайте файл .env с BOT_TOKEN=ваш_токен")
        return
    
    # Создаем приложение
    application = Application.builder().token(config.BOT_TOKEN).build()
    
    # Регистрируем обработчики команд
    application.add_handler(CommandHandler("start", start))
    application.add_handler(CommandHandler("add", add_sneeze))
    application.add_handler(CommandHandler("stats", show_stats))
    application.add_handler(CommandHandler("chart", show_chart))
    application.add_handler(CommandHandler("edit", edit_date))
    application.add_handler(CommandHandler("today", show_today))
    
    # Админ-команды
    application.add_handler(CommandHandler("admin_stats", admin_stats))
    application.add_handler(CommandHandler("admin_export", admin_export))
    
    # Обработчик текстовых сообщений (для записи чисел и кнопки "Чихнуть")
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_number_message))
    
    # Запускаем бота
    logger.info("Бот запущен...")
    application.run_polling(allowed_updates=Update.ALL_TYPES)


if __name__ == '__main__':
    main()
